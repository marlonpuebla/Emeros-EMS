#!/usr/bin/env python3
"""Import employees from Excel roster into HR database."""
import sys, json, sqlite3, datetime, calendar
from openpyxl import load_workbook

def fix_date(val):
    if val is None: return None
    if isinstance(val, datetime.datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, str):
        val = val.strip()
        if not val or val.upper() == 'N/A': return None
        import re
        m = re.match(r'(\d{1,2})/(\d{1,2})/(\d{4})', val)
        if m:
            month, day, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
            max_day = calendar.monthrange(year, month)[1]
            day = min(day, max_day)
            return f"{year:04d}-{month:02d}-{day:02d}"
    return None

def clean_str(val):
    if val is None: return None
    s = str(val).strip().replace('\xa0', '').strip()
    return None if (not s or s.upper() == 'N/A') else s

def clean_num(val):
    if val is None: return None
    if isinstance(val, float) and val == int(val): val = int(val)
    s = str(val).strip()
    return None if (not s or s.upper() == 'N/A') else s

if __name__ == '__main__':
    xlsx_path = sys.argv[1]
    db_path   = sys.argv[2]
    mode      = sys.argv[3] if len(sys.argv) > 3 else 'merge'  # merge or replace

    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))[1:]

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    if mode == 'replace':
        c.execute("DELETE FROM employees")

    inserted = updated = skipped = 0
    warnings = []

    for i, row in enumerate(rows, start=2):
        if len(row) < 10 or not row[1]: continue
        try:
            emp = {
                'status': (row[0] or 'ACTIVE').strip().lower(),
                'last_name': clean_str(row[1]),
                'first_name': clean_str(row[2]),
                'employment_date': fix_date(row[3]),
                'dob': fix_date(row[4]),
                'gender': clean_str(row[5]),
                'phone': clean_num(row[6]),
                'email': clean_str(row[7]),
                'position_ahca': clean_str(row[8]),
                'position_om': clean_str(row[9]),
                'emergency_contact': clean_str(row[10]) if len(row) > 10 else None,
                'emergency_contact_phone': clean_num(row[11]) if len(row) > 11 else None,
                'license_number': clean_str(row[12]) if len(row) > 12 else None,
                'license_expiration': fix_date(row[13]) if len(row) > 13 else None,
                'caqh': clean_num(row[14]) if len(row) > 14 else None,
                'npi': clean_num(row[15]) if len(row) > 15 else None,
                'taxonomy': clean_str(row[16]) if len(row) > 16 else None,
                'medicare': clean_str(row[17]) if len(row) > 17 else None,
                'medicaid': clean_num(row[18]) if len(row) > 18 else None,
                'dea': clean_str(row[19]) if len(row) > 19 else None,
                'dea_expiration': fix_date(row[20]) if len(row) > 20 else None,
                'sunbiz_co': clean_str(row[21]) if len(row) > 21 else None,
                'ein': clean_num(row[22]) if len(row) > 22 else None,
                'driver_license': clean_str(row[23]) if len(row) > 23 else None,
                'driver_license_expiration': fix_date(row[24]) if len(row) > 24 else None,
                'ssn': clean_num(row[25]) if len(row) > 25 else None,
                'ahca_background_expiration': fix_date(row[26]) if len(row) > 26 else None,
                'professional_liability_expiration': fix_date(row[27]) if len(row) > 27 else None,
                'ceu_expiration': fix_date(row[28]) if len(row) > 28 else None,
                'cpr_bls_expiration': fix_date(row[29]) if len(row) > 29 else None,
                'passport_expiration': fix_date(row[30]) if len(row) > 30 else None,
                'e_verified': fix_date(row[31]) if isinstance(row[31], datetime.datetime) else clean_str(row[31]) if len(row) > 31 else None,
                'exemption_worker_comp_expiration': fix_date(row[32]) if len(row) > 32 else None,
                'yearly_evaluation_due': fix_date(row[33]) if len(row) > 33 else None,
                'rehired_date': fix_date(row[34]) if len(row) > 34 else None,
                'termination_date': fix_date(row[35]) if len(row) > 35 else None,
            }
            if not emp['last_name']:
                skipped += 1; continue

            # Try to find existing by name+dob
            existing = None
            if emp['dob']:
                c.execute("SELECT id FROM employees WHERE last_name=? AND first_name=? AND dob=?",
                          (emp['last_name'], emp['first_name'], emp['dob']))
                existing = c.fetchone()

            if existing and mode == 'merge':
                sets = ', '.join([f"{k}=?" for k in emp.keys()])
                c.execute(f"UPDATE employees SET {sets} WHERE id=?",
                          list(emp.values()) + [existing['id']])
                updated += 1
            else:
                cols = ', '.join(emp.keys())
                phs  = ', '.join(['?'] * len(emp))
                c.execute(f"INSERT INTO employees ({cols}) VALUES ({phs})", list(emp.values()))
                inserted += 1
        except Exception as e:
            warnings.append(f"Row {i}: {str(e)}")

    conn.commit()
    conn.close()
    print(json.dumps({'ok': True, 'inserted': inserted, 'updated': updated,
                      'skipped': skipped, 'warnings': warnings}))
